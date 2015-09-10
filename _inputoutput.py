import codecs
import translitcodec
import openpyxl
import csv
import datetime
import pytz
import os
import requests
import urllib2
import json
import time
import mechanize
from bs4 import BeautifulSoup




def collectStockData(cid, startDt, endDt, path):
	"""
	Returns the OHLCV stock data of a stock in given time period.

	Cid refers to the stock cid parameter in Google Finance URls, which
	is unique for each stock.
	"""
	stockData = []

	#if the file already exists, just load from that
	if os.path.isfile(path):
		stockData = read(path)
		#convert prices from strings to floats
		for line in stockData:
			line[1] = float(line[1])
			line[2] = float(line[2])
			line[3] = float(line[3])
			line[4] = float(line[4])
			line[5] = float(line[5])
		return stockData
	#elsewise, load from Google Finance
	else:

		startDtStr = datetime.datetime.strftime(startDt, "%m/%d/%y")
		endDtStr = datetime.datetime.strftime(endDt, "%m/%d/%y")

		query = "http://www.google.com/finance/historical?cid="+str(cid)+"&startdate="+startDtStr+"&enddate="+endDtStr+"&output=csv"

		rawStockData = requests.get(query)
		print(rawStockData.content)

		lines = rawStockData.content.split("\n")
		for line in lines[1:-1]:
			cells = line.split(",")
			#add padding to day and year so %d in strptime will work
			date = cells[0].split("-")
			if len(date) == 3:
				pad = "0" if len(date[0]) == 1 else ""
				date[0] = pad+date[0]
				pad = "0" if len(date[2]) == 1 else ""
				date[2] = pad+date[2]
				cells[0] = date[0]+"-"+date[1]+"-"+date[2]
			#create datetime object for data date
			dt = datetime.datetime.strptime(cells[0], "%d-%b-%y")
			o = float(cells[1])
			h = float(cells[2])
			l = float(cells[3])
			c = float(cells[4])
			#after certain period, google finance stops delivering volume
			#info, so just use -1 as placeholder
			v = -1 if cells[5] == "-" else float(cells[5])

			stockData.append([dt, o, h, l, c, v])

	#the data is originally final->initial, but it must be in
	#form initial->final
	stockData = list(reversed(stockData))

	return stockData




def _dayIter(start, end, delta):
	"""
	Returns datetimes from [start, end) for use in iteration.
	"""
	current = start
	while current < end:
		yield current
		current += delta




def collectArticles(terms, startDt, endDt):
	iters = _dayIter(startDt, endDt, datetime.timedelta(days=1))

	browser = mechanize.Browser()
	browser.set_handle_robots(False)
	browser.addheaders = [("User-agent", "Mozilla/5.0 (X11; Linux x86_64) Chrome/43.0.2357.132 AppleWebKit/535.21 (KHTML, like Gecko) Safari/535.21")]
	#additional browser options
	#Mozilla/5.0 (X11; Linux x86_64) Chrome/43.0.2357.132 AppleWebKit/535.21 (KHTML, like Gecko) Safari/535.21


	urls = []
	for date in iters:
		day = date.day
		month = date.month
		year = date.year

		queryBase = "https://www.google.com/search?q="
		for term in terms:
			queryBase += term+"+"
		queryBase = queryBase[:-1]
		query = queryBase + "&safe=off&tbs=cdr%3A1%2Ccd_min%3A"+str(month)+"%2F"+str(day)+"%2F"+str(year)+"%2Ccd_max%3A"+str(month)+"%2F"+str(day)+"%2F"+str(year)+"&tbm=nws"

		try:
			browser.open(query)
			html = browser.response().read()
			page = BeautifulSoup(html)

			items = page.find_all("li", {"class" : "g"})
			for item in items:
				url = item.find("a", {"class" : "l _HId"}, href=True)
				source = item.find("span", {"class" : "_tQb _IId"})
				line = [source.text, url["href"]]
				print(url["href"])
				urls.append(line)

			time.sleep(2)
		except:
			continue


	finalUrls = []
	EST = pytz.timezone("America/New_York")

	for i, line in enumerate(urls):
		try:
			queryBase = "http://api.diffbot.com/v3/article?token=ea0342bf7d2d09c8f0248ac7665a123d&url="
			query = queryBase + line[1]
			print("scraping: "+line[1])

			response = urllib2.urlopen(query)
			data = json.loads(response.read())

			if "error" in data:
				print("Error: Diffbot was unable to scrape this webpage.")
				continue
				

			text = data["objects"][0]["text"]
			try: #preffered, but sometimes breaks
				text = codecs.encode(text, "translit/long/ascii")
			except:
				text = text.encode("ascii", "ignore")

			#if date is not found page, simply take the date of the previous entry
			if "date" not in data["objects"][0]:
				print("Error: Date not found from Diffbot data, so using dt of previous article.")
				if i != 1:
					dt = finalUrls[len(finalUrls)-1][0]
				else:	#if this is the first url searched, just give it the startDt
					dt = startDt.replace(tzinfo=EST)

			else:
				dt = data["objects"][0]["date"]
				dt = dt[:-4]
				dt = datetime.datetime.strptime(dt, "%a, %d %b %Y %H:%M:%S")
				if dt.hour == 0 and dt.min == 0 and dt.sec == 0:
					dt = dt.replace(tzinfo=EST)
					print(str(dt)+": 00:00:00 GMT time replaced to EST/EDT")
				else:
					dt = dt.replace(tzinfo=pytz.utc)
					dt = EST.normalize(dt.astimezone(EST))
					print(str(dt)+": GMT converted properly to EST/EDT")

			source = line[0]
			url = line[1]

			newLine = [dt, source, url, text]
			finalUrls.append(newLine)

		except:
			print("Error has occurred.")

	return finalUrls



def write(path, data):
	"""
	Writes data to file, line by line.

	Writes the article data to a .csv file in format:
		"dateTime, url"
	Writes the stock data to a .csv file in format:
		"dateTime, o, h, l, c, v"
	"""
	file = open(path, "wb")
	writer = csv.writer(file, delimiter=",")
	for line in data:
		#print line[0]
		#print line[1]
		newLine = []
		newLine.append(line[0].strftime("%Y-%m-%d %H:%M:%S").encode("ascii", "ignore"))
		for i in range(1, len(line)):
			newLine.append(line[i])
		writer.writerow(newLine)
	file.close()



def read(path):
	"""
	Reads data--with firstCol=dateTime, and then other data--into array.
	"""
	data = []
	file = open(path)
	reader = csv.reader(file, delimiter=",")
	EST = pytz.timezone("America/New_York")
	for line in reader:
		newLine = []
		dt = datetime.datetime.strptime(line[0], "%Y-%m-%d %H:%M:%S")
		#dt = dt.replace(tzinfo=EST) #DEBUG FOR TESTSTRATEGY(), RESTORE LATER
		newLine.append(dt)
		for i in range(1, len(line)):
			newLine.append(float(line[i]))
		data.append(newLine)
	file.close()
	return data

def readRaw(path):
	"""
	Reads data--with firstCol=dateTime, and then other data--into array.
	"""
	data = []
	file = open(path)
	reader = csv.reader(file, delimiter=",")
	EST = pytz.timezone("America/New_York")
	for line in reader:
		newLine = []
		dt = datetime.datetime.strptime(line[0], "%Y-%m-%d %H:%M:%S")
		dt = dt.replace(tzinfo=EST)
		newLine.append(dt)
		for i in range(1, len(line)):
			newLine.append(line[i])
		data.append(newLine)
	file.close()
	return data


def getSents(path):
	"""
	Gets the sentiment data from Semantria excel sheet.
	"""
	wb = openpyxl.load_workbook(path)
	ws = wb["artSent_Detailed"]

	#collect all sentiments from column E (many duplicates)
	sents = []
	for col in ws.iter_rows("E2:E"+str(len(ws.rows))):
		for cell in col:
			sents.append(cell.value)

	#collect corresponding, unique, id of each sentiment
	ids = []
	for col in ws.iter_rows("A2:A"+str(len(ws.rows))):
		for cell in col:
				ids.append(cell.value)

	#remove duplicates
	prev_id = None
	consol_sents = []
	for i in range(len(ids)):
		if ids[i] != prev_id:
			consol_sents.append(sents[i])
			prev_id = ids[i]

	for i,s in enumerate(consol_sents):
		consol_sents[i] = float(s)

	return consol_sents

def getSentsNER(path):
	"""
	Gets the sentiment data from Semantria excel sheet.
	Only gets the "Amazon.com" NER sentiment.
	"""
	wb = openpyxl.load_workbook(path)
	ws = wb["artSentNER"]

	sents = []
	dts = []
	EST = pytz.timezone("America/New_York")
	#collect corresponding datetime of each sentiment
	for col in ws.iter_rows("A1:A"+str(len(ws.rows))):
		for cell in col:
			dt = cell.value.replace(tzinfo=EST)
			dts.append(dt)
	#corresponding sent values:
	for col in ws.iter_rows("B1:B"+str(len(ws.rows))):
		for cell in col:
			sents.append(float(cell.value))

	data = [dts, sents]
	data = zip(*data)

	for line in data:
		print line

	return data


def getAllArtData(path, entSent=False):
	"""
	Gets the Wikipedia category data from Semantria excel sheet.

	It's in format: [dt, site, url, text, sent, [ents], ([entsent]), [themes], [wikiCats], [userCats]]
	"""
	wb = openpyxl.load_workbook(path)
	ws = wb["artSent_Detailed"]
	ws2 = wb["rawArticleData"]
	EST = pytz.timezone("America/New_York")

	#collect basic data:
	#datetime
	dts = []
	for col in ws2.iter_rows("A1:A"+str(len(ws2.rows))):
		for cell in col:
			dt = cell.value.replace(tzinfo=EST)
			dts.append(dt)
	#site
	sites = []
	for col in ws2.iter_rows("B1:B"+str(len(ws2.rows))):
		for cell in col:
			sites.append(cell.value)
	#url
	urls = []
	for col in ws2.iter_rows("C1:C"+str(len(ws2.rows))):
		for cell in col:
			urls.append(cell.value)
	#text
	texts = []
	for col in ws2.iter_rows("D1:D"+str(len(ws2.rows))):
		for cell in col:
			texts.append(cell.value)

	basicData = [list(z) for z in zip(dts, sites, urls, texts)]



	#collect analysis data:
	#ids
	ids = []
	for col in ws.iter_rows("A2:A"+str(len(ws.rows))):
		for cell in col:
			ids.append(cell.value)
	#collect all sentiments
	sents = []
	for col in ws.iter_rows("E2:E"+str(len(ws.rows))):
		for cell in col:
			sents.append(cell.value)
	#entities
	ents = []
	for col in ws.iter_rows("I2:I"+str(len(ws.rows))):
		for cell in col:
			ents.append(cell.value)
	#if we also want the entity sentiments, then return ents and corresponding
	#sents as a list of lists
	if entSent:
		entSents = []
		for col in ws.iter_rows("K2:K"+str(len(ws.rows))):
			for cell in col:
				entSents.append(cell.value)
		ent2D = []
		for i in range(len(ents)):
			ent2D.append([ents[i], entSents[i]])
		ents = ent2D

	#themes
	themes = []
	for col in ws.iter_rows("N2:N"+str(len(ws.rows))):
		for cell in col:
			themes.append(cell.value)
	#category
	cats = []
	for col in ws.iter_rows("R2:R"+str(len(ws.rows))):
		for cell in col:
			cats.append(cell.value)
	#subcategory
	subCats = []
	for col in ws.iter_rows("S2:S"+str(len(ws.rows))):
		for cell in col:
			subCats.append(cell.value)
	#user category
	userCats = []
	for col in ws.iter_rows("U2:U"+str(len(ws.rows))):
		for cell in col:
			userCats.append(cell.value)


	#reformat data
	anlData = []
	chunk = []
	prevId = ids[0]
	for i in range(len(ids)):
		if ids[i] != prevId:
			anlData.append(chunkExcel(chunk, entSent))
			chunk = []
		chunk.append([sents[i], ents[i], themes[i], cats[i], subCats[i], userCats[i]])
		prevId = ids[i]
	anlData.append(chunkExcel(chunk, entSent))	#final article is not caught

	data = []
	for i in range(len(basicData)):
		line = basicData[i] + anlData[i]
		data.append(line)
	return data



def chunkExcel(chunk, entSent=False):
	data = []

	ents = []
	entSents = []
	themes = []
	cats = []
	userCats = []
	for line in chunk:
		if line[1] != None:
			if entSent:
				ents.append(line[1][0])
				entSents.append(line[1][1])
			else:
				ents.append(line[1])
		if line[2] != None:
			themes.append(line[2])
		if line[4] != None:
			cats.append(str(line[3])+":"+str(line[4]))
		if line[5] != None:
			userCats.append(line[5])
	
	if entSent:
		data = [chunk[0][0], ents, entSents, themes, cats, userCats]
	else:
		data = [chunk[0][0], ents, themes, cats, userCats]

	return data